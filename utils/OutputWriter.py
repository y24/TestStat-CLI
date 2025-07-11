import csv
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class OutputWriter:
    """CSV/Excel形式での出力を担当するクラス"""
    
    def __init__(self, verbose_logger=None):
        self.verbose_logger = verbose_logger
    
    def write_csv(self, data, output_file, is_multiple_files=False):
        """CSV形式でファイルに出力"""
        try:
            # 出力ディレクトリの確認・作成
            output_dir = os.path.dirname(output_file)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)
                if self.verbose_logger:
                    self.verbose_logger.log(f"出力ディレクトリを作成しました: {output_dir}")
            
            # ファイルが既に存在する場合の確認
            if os.path.exists(output_file):
                if self.verbose_logger:
                    self.verbose_logger.log(f"既存ファイルを上書きします: {output_file}")
            
            with open(output_file, 'w', newline='', encoding='utf-8-sig') as csvfile:
                if is_multiple_files:
                    self._write_multiple_files_csv(data, csvfile)
                else:
                    self._write_single_file_csv(data, csvfile)
            
            if self.verbose_logger:
                self.verbose_logger.log(f"CSVファイルを出力しました: {output_file}")
            
            return True, None
            
        except PermissionError:
            error_msg = f"ファイルの書き込み権限がありません: {output_file}"
            if self.verbose_logger:
                self.verbose_logger.log(f"ERROR: {error_msg}")
            return False, error_msg
        except OSError as e:
            error_msg = f"ファイル出力エラー: {output_file}, 詳細: {e}"
            if self.verbose_logger:
                self.verbose_logger.log(f"ERROR: {error_msg}")
            return False, error_msg
        except Exception as e:
            error_msg = f"予期しないエラーが発生しました: {e}"
            if self.verbose_logger:
                self.verbose_logger.log(f"ERROR: {error_msg}")
            return False, error_msg
    
    def _write_single_file_csv(self, data, csvfile):
        """単一ファイルのCSV出力"""
        writer = csv.writer(csvfile)
        
        # TOTAL RESULTS
        if "total" in data:
            writer.writerow(["Category", "Pass", "Fixed", "Fail", "Blocked", "Suspend", "N/A", "Total", "完了数", "消化数", "完了率(%)", "消化率(%)"])
            total = data["total"]
            writer.writerow([
                "TOTAL RESULTS",
                total.get("Pass", 0),
                total.get("Fixed", 0),
                total.get("Fail", 0),
                total.get("Blocked", 0),
                total.get("Suspend", 0),
                total.get("N/A", 0),
                total.get("Total", 0),
                total.get("完了数", 0),
                total.get("消化数", 0),
                total.get("完了率(%)", 0),
                total.get("消化率(%)", 0)
            ])
            writer.writerow([])  # 空行
        
        # STATISTICS
        if "stats" in data:
            writer.writerow(["Metric", "Count"])
            stats = data["stats"]
            for metric in ["all", "available", "executed", "completed", "incompleted", "planned"]:
                if metric in stats:
                    writer.writerow([metric.capitalize(), stats[metric]])
            writer.writerow([])  # 空行
        
        # DAILY BREAKDOWN
        if "daily" in data:
            writer.writerow(["Date", "Pass", "Fixed", "Fail", "Blocked", "Suspend", "N/A", "完了数", "消化数", "計画数"])
            for date, daily_data in sorted(data["daily"].items()):
                writer.writerow([
                    date,
                    daily_data.get("Pass", 0),
                    daily_data.get("Fixed", 0),
                    daily_data.get("Fail", 0),
                    daily_data.get("Blocked", 0),
                    daily_data.get("Suspend", 0),
                    daily_data.get("N/A", 0),
                    daily_data.get("完了数", 0),
                    daily_data.get("消化数", 0),
                    daily_data.get("計画数", 0)
                ])
            writer.writerow([])  # 空行
        
        # BY NAME
        if "by_name" in data:
            # 担当者名を取得
            testers = set()
            for daily_data in data["by_name"].values():
                testers.update(daily_data.keys())
            testers = sorted([a for a in testers if a not in ["完了数", "消化数", "計画数"]])
            
            if testers:
                headers = ["Date"] + testers + ["完了数", "消化数", "計画数"]
                writer.writerow(headers)
                for date, name_data in sorted(data["by_name"].items()):
                    row = [date]
                    for tester in testers:
                        row.append(name_data.get(tester, 0))
                    row.extend([
                        name_data.get("完了数", 0),
                        name_data.get("消化数", 0),
                        name_data.get("計画数", 0)
                    ])
                    writer.writerow(row)
                writer.writerow([])  # 空行
        
        # BY ENVIRONMENT
        if "by_env" in data:
            writer.writerow(["Environment", "Date", "Pass", "Fixed", "Fail", "Blocked", "Suspend", "N/A", "完了数", "消化数", "計画数"])
            for env_name, env_data in sorted(data["by_env"].items()):
                for date, env_stats in sorted(env_data.items()):
                    writer.writerow([
                        env_name,
                        date,
                        env_stats.get("Pass", 0),
                        env_stats.get("Fixed", 0),
                        env_stats.get("Fail", 0),
                        env_stats.get("Blocked", 0),
                        env_stats.get("Suspend", 0),
                        env_stats.get("N/A", 0),
                        env_stats.get("完了数", 0),
                        env_stats.get("消化数", 0),
                        env_stats.get("計画数", 0)
                    ])
    
    def _write_multiple_files_csv(self, data, csvfile):
        """複数ファイルのCSV出力"""
        writer = csv.writer(csvfile)
        
        # SUMMARY TOTAL RESULTS
        if "summary" in data and "total_results" in data["summary"]:
            writer.writerow(["Category", "Pass", "Fixed", "Fail", "Blocked", "Suspend", "N/A", "Total", "完了数", "消化数", "完了率(%)", "消化率(%)"])
            total = data["summary"]["total_results"]
            writer.writerow([
                "SUMMARY TOTAL RESULTS",
                total.get("Pass", 0),
                total.get("Fixed", 0),
                total.get("Fail", 0),
                total.get("Blocked", 0),
                total.get("Suspend", 0),
                total.get("N/A", 0),
                total.get("Total", 0),
                total.get("完了数", 0),
                total.get("消化数", 0),
                total.get("完了率(%)", 0),
                total.get("消化率(%)", 0)
            ])
            writer.writerow([])  # 空行
        
        # SUMMARY STATISTICS
        if "summary" in data and "total_stats" in data["summary"]:
            writer.writerow(["Metric", "Total"])
            stats = data["summary"]["total_stats"]
            for metric in ["all", "available", "executed", "completed", "incompleted", "planned"]:
                if metric in stats:
                    writer.writerow([f"{metric.capitalize()} Cases", stats[metric]])
            writer.writerow([])  # 空行
        
        # INDIVIDUAL FILES
        if "files" in data:
            writer.writerow(["File", "Total Cases", "Available Cases", "Excluded Cases", "Pass", "Fixed", "Fail", "Blocked", "Suspend", "N/A", "Total", "完了数", "消化数", "完了率(%)", "消化率(%)"])
            for file_data in data["files"]:
                if "total" in file_data and "stats" in file_data:
                    total = file_data["total"]
                    stats = file_data["stats"]
                    writer.writerow([
                        file_data.get("file", ""),
                        stats.get("all", 0),
                        stats.get("available", 0),
                        stats.get("excluded", 0),
                        total.get("Pass", 0),
                        total.get("Fixed", 0),
                        total.get("Fail", 0),
                        total.get("Blocked", 0),
                        total.get("Suspend", 0),
                        total.get("N/A", 0),
                        total.get("Total", 0),
                        total.get("完了数", 0),
                        total.get("消化数", 0),
                        total.get("完了率(%)", 0),
                        total.get("消化率(%)", 0)
                    ])
        
        # BY ENVIRONMENT
        if "by_env" in data:
            writer.writerow(["Environment", "Date", "Pass", "Fixed", "Fail", "Blocked", "Suspend", "N/A", "完了数", "消化数", "計画数"])
            for env_name, env_data in sorted(data["by_env"].items()):
                for date, env_stats in sorted(env_data.items()):
                    writer.writerow([
                        env_name,
                        date,
                        env_stats.get("Pass", 0),
                        env_stats.get("Fixed", 0),
                        env_stats.get("Fail", 0),
                        env_stats.get("Blocked", 0),
                        env_stats.get("Suspend", 0),
                        env_stats.get("N/A", 0),
                        env_stats.get("完了数", 0),
                        env_stats.get("消化数", 0),
                        env_stats.get("計画数", 0)
                    ])
    
    def write_excel(self, data, output_file, is_multiple_files=False, filters=None):
        """Excel形式でファイルに出力"""
        try:
            # 出力ディレクトリの確認・作成
            output_dir = os.path.dirname(output_file)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)
                if self.verbose_logger:
                    self.verbose_logger.log(f"出力ディレクトリを作成しました: {output_dir}")
            
            # ファイルが既に存在する場合の確認
            if os.path.exists(output_file):
                if self.verbose_logger:
                    self.verbose_logger.log(f"既存ファイルを上書きします: {output_file}")
            
            # ワークブック作成
            wb = Workbook()
            
            # デフォルトシートを削除
            wb.remove(wb.active)
            
            if is_multiple_files:
                self._write_multiple_files_excel(data, wb, filters)
            else:
                self._write_single_file_excel(data, wb, filters)
            
            # ファイル保存
            wb.save(output_file)
            
            if self.verbose_logger:
                self.verbose_logger.log(f"Excelファイルを出力しました: {output_file}")
            
            return True, None
            
        except PermissionError:
            error_msg = f"ファイルの書き込み権限がありません: {output_file}"
            if self.verbose_logger:
                self.verbose_logger.log(f"ERROR: {error_msg}")
            return False, error_msg
        except OSError as e:
            error_msg = f"ファイル出力エラー: {output_file}, 詳細: {e}"
            if self.verbose_logger:
                self.verbose_logger.log(f"ERROR: {error_msg}")
            return False, error_msg
        except Exception as e:
            error_msg = f"予期しないエラーが発生しました: {e}"
            if self.verbose_logger:
                self.verbose_logger.log(f"ERROR: {error_msg}")
            return False, error_msg
    
    def _write_single_file_excel(self, data, wb, filters=None):
        """単一ファイルのExcel出力"""
        # TOTAL RESULTS シート
        ws_total = wb.create_sheet("TOTAL RESULTS")
        self._write_total_results_sheet(ws_total, data.get("total", {}))
        
        # STATISTICS シート
        ws_stats = wb.create_sheet("STATISTICS")
        self._write_statistics_sheet(ws_stats, data.get("stats", {}))
        
        # DAILY BREAKDOWN シート
        if "daily" in data:
            ws_daily = wb.create_sheet("DAILY BREAKDOWN")
            self._write_daily_breakdown_sheet(ws_daily, data["daily"])
        
        # BY NAME シート
        if "by_name" in data:
            ws_name = wb.create_sheet("BY NAME")
            self._write_by_name_sheet(ws_name, data["by_name"])
        
        # BY ENVIRONMENT シート
        if "by_env" in data:
            ws_env = wb.create_sheet("BY ENVIRONMENT")
            self._write_by_environment_sheet(ws_env, data["by_env"])
        
        # METADATA シート
        ws_meta = wb.create_sheet("METADATA")
        self._write_metadata_sheet(ws_meta, data, filters, is_multiple=False)
    
    def _write_multiple_files_excel(self, data, wb, filters=None):
        """複数ファイルのExcel出力"""
        # SUMMARY TOTAL RESULTS シート
        ws_summary = wb.create_sheet("SUMMARY TOTAL RESULTS")
        if "summary" in data and "total_results" in data["summary"]:
            self._write_total_results_sheet(ws_summary, data["summary"]["total_results"])
        
        # SUMMARY STATISTICS シート
        ws_stats = wb.create_sheet("SUMMARY STATISTICS")
        if "summary" in data and "total_stats" in data["summary"]:
            self._write_statistics_sheet(ws_stats, data["summary"]["total_stats"], is_summary=True)
        
        # INDIVIDUAL FILES シート
        if "files" in data:
            ws_files = wb.create_sheet("INDIVIDUAL FILES")
            self._write_individual_files_sheet(ws_files, data["files"])
        
        # DAILY BREAKDOWN シート
        if "files" in data:
            ws_daily = wb.create_sheet("DAILY BREAKDOWN")
            self._write_combined_daily_sheet(ws_daily, data["files"])
        
        # BY NAME シート
        if "files" in data:
            ws_name = wb.create_sheet("BY NAME")
            self._write_combined_by_name_sheet(ws_name, data["files"])
        
        # BY ENVIRONMENT シート
        if "files" in data:
            ws_env = wb.create_sheet("BY ENVIRONMENT")
            self._write_combined_by_environment_sheet(ws_env, data["files"])
        
        # METADATA シート
        ws_meta = wb.create_sheet("METADATA")
        self._write_metadata_sheet(ws_meta, data, filters, is_multiple=True)
    
    def _write_total_results_sheet(self, ws, total_data):
        """TOTAL RESULTS シートの書き込み"""
        headers = ["Category", "Pass", "Fixed", "Fail", "Blocked", "Suspend", "N/A", "Total", "完了数", "消化数", "完了率(%)", "消化率(%)"]
        data = [
            "TOTAL RESULTS",
            total_data.get("Pass", 0),
            total_data.get("Fixed", 0),
            total_data.get("Fail", 0),
            total_data.get("Blocked", 0),
            total_data.get("Suspend", 0),
            total_data.get("N/A", 0),
            total_data.get("Total", 0),
            total_data.get("完了数", 0),
            total_data.get("消化数", 0),
            total_data.get("完了率(%)", 0),
            total_data.get("消化率(%)", 0)
        ]
        
        self._write_sheet_with_formatting(ws, [headers, data])
    
    def _write_statistics_sheet(self, ws, stats_data, is_summary=False):
        """STATISTICS シートの書き込み"""
        if is_summary:
            headers = ["Metric", "Total"]
            data = []
            for metric in ["all", "available", "executed", "completed", "incompleted", "planned"]:
                if metric in stats_data:
                    data.append([f"{metric.capitalize()} Cases", stats_data[metric]])
        else:
            headers = ["Metric", "Count"]
            data = []
            for metric in ["all", "available", "executed", "completed", "incompleted", "planned"]:
                if metric in stats_data:
                    data.append([metric.capitalize(), stats_data[metric]])
        
        self._write_sheet_with_formatting(ws, [headers] + data)
    
    def _write_daily_breakdown_sheet(self, ws, daily_data):
        """DAILY BREAKDOWN シートの書き込み"""
        headers = ["Date", "Pass", "Fixed", "Fail", "Blocked", "Suspend", "N/A", "完了数", "消化数", "計画数"]
        data = []
        for date, daily_stats in sorted(daily_data.items()):
            data.append([
                date,
                daily_stats.get("Pass", 0),
                daily_stats.get("Fixed", 0),
                daily_stats.get("Fail", 0),
                daily_stats.get("Blocked", 0),
                daily_stats.get("Suspend", 0),
                daily_stats.get("N/A", 0),
                daily_stats.get("完了数", 0),
                daily_stats.get("消化数", 0),
                daily_stats.get("計画数", 0)
            ])
        
        self._write_sheet_with_formatting(ws, [headers] + data)
    
    def _write_by_name_sheet(self, ws, by_name_data):
        """BY NAME シートの書き込み"""
        # 担当者名を取得
        testers = set()
        for daily_data in by_name_data.values():
            testers.update(daily_data.keys())
        testers = sorted([a for a in testers if a not in ["完了数", "消化数", "計画数"]])
        
        if testers:
            headers = ["Date"] + testers + ["完了数", "消化数", "計画数"]
            data = []
            for date, name_data in sorted(by_name_data.items()):
                row = [date]
                for tester in testers:
                    row.append(name_data.get(tester, 0))
                row.extend([
                    name_data.get("完了数", 0),
                    name_data.get("消化数", 0),
                    name_data.get("計画数", 0)
                ])
                data.append(row)
            
            self._write_sheet_with_formatting(ws, [headers] + data)
    
    def _write_by_environment_sheet(self, ws, by_env_data):
        """BY ENVIRONMENT シートの書き込み"""
        headers = ["Environment", "Date", "Pass", "Fixed", "Fail", "Blocked", "Suspend", "N/A", "完了数", "消化数", "計画数"]
        data = []
        for env_name, env_data in sorted(by_env_data.items()):
            for date, env_stats in sorted(env_data.items()):
                data.append([
                    env_name,
                    date,
                    env_stats.get("Pass", 0),
                    env_stats.get("Fixed", 0),
                    env_stats.get("Fail", 0),
                    env_stats.get("Blocked", 0),
                    env_stats.get("Suspend", 0),
                    env_stats.get("N/A", 0),
                    env_stats.get("完了数", 0),
                    env_stats.get("消化数", 0),
                    env_stats.get("計画数", 0)
                ])
        self._write_sheet_with_formatting(ws, [headers] + data)

    def _write_individual_files_sheet(self, ws, files_data):
        """INDIVIDUAL FILES シートの書き込み"""
        headers = ["File", "Total Cases", "Available Cases", "Excluded Cases", "Pass", "Fixed", "Fail", "Blocked", "Suspend", "N/A", "Total", "完了数", "消化数", "完了率(%)", "消化率(%)"]
        data = []
        for file_data in files_data:
            if "total" in file_data and "stats" in file_data:
                total = file_data["total"]
                stats = file_data["stats"]
                data.append([
                    file_data.get("file", ""),
                    stats.get("all", 0),
                    stats.get("available", 0),
                    stats.get("excluded", 0),
                    total.get("Pass", 0),
                    total.get("Fixed", 0),
                    total.get("Fail", 0),
                    total.get("Blocked", 0),
                    total.get("Suspend", 0),
                    total.get("N/A", 0),
                    total.get("Total", 0),
                    total.get("完了数", 0),
                    total.get("消化数", 0),
                    total.get("完了率(%)", 0),
                    total.get("消化率(%)", 0)
                ])
        
        self._write_sheet_with_formatting(ws, [headers] + data)
    
    def _write_combined_daily_sheet(self, ws, files_data):
        """統合DAILY BREAKDOWN シートの書き込み"""
        # 全ファイルの日別データを統合
        combined_daily = {}
        for file_data in files_data:
            if "daily" in file_data:
                for date, daily_stats in file_data["daily"].items():
                    if date not in combined_daily:
                        combined_daily[date] = {
                            "Pass": 0, "Fixed": 0, "Fail": 0, "Blocked": 0, "Suspend": 0, "N/A": 0,
                            "完了数": 0, "消化数": 0, "計画数": 0
                        }
                    for key, value in daily_stats.items():
                        if key in combined_daily[date]:
                            combined_daily[date][key] += value
        
        self._write_daily_breakdown_sheet(ws, combined_daily)
    
    def _write_combined_by_name_sheet(self, ws, files_data):
        """統合BY NAME シートの書き込み"""
        # 全ファイルの担当者別データを統合
        combined_by_name = {}
        for file_data in files_data:
            if "by_name" in file_data:
                for date, name_data in file_data["by_name"].items():
                    if date not in combined_by_name:
                        combined_by_name[date] = {}
                    for name, count in name_data.items():
                        if name not in combined_by_name[date]:
                            combined_by_name[date][name] = 0
                        combined_by_name[date][name] += count
        
        self._write_by_name_sheet(ws, combined_by_name)
    
    def _write_combined_by_environment_sheet(self, ws, files_data):
        """統合BY ENVIRONMENT シートの書き込み"""
        # 全ファイルの環境別データを統合
        combined_by_env = {}
        for file_data in files_data:
            if "by_env" in file_data:
                for date, env_data in file_data["by_env"].items():
                    if date not in combined_by_env:
                        combined_by_env[date] = {}
                    for env_name, env_stats in env_data.items():
                        if env_name not in combined_by_env[date]:
                            combined_by_env[date][env_name] = {
                                "Pass": 0, "Fixed": 0, "Fail": 0, "Blocked": 0, "Suspend": 0, "N/A": 0,
                                "完了数": 0, "消化数": 0, "計画数": 0
                            }
                        for key, value in env_stats.items():
                            if key in combined_by_env[date][env_name]:
                                combined_by_env[date][env_name][key] += value
        
        self._write_by_environment_sheet(ws, combined_by_env)
    
    def _write_metadata_sheet(self, ws, data, filters=None, is_multiple=False):
        """METADATA シートの書き込み"""
        metadata = [
            ["項目", "値"],
            ["処理日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["出力形式", "Excel"]
        ]
        
        if is_multiple and "summary" in data:
            metadata.extend([
                ["処理ファイル数", data["summary"].get("processed_files", 0)],
                ["総処理時間", f"{data['summary'].get('processing_time', 0)}秒"]
            ])
        
        if filters:
            filter_info = []
            if "date_range" in filters:
                date_range = filters["date_range"]
                if date_range.get("start") and date_range.get("end"):
                    filter_info.append(f"日付範囲: {date_range['start']} から {date_range['end']}")
                elif date_range.get("start"):
                    filter_info.append(f"日付範囲: {date_range['start']} 以降")
                elif date_range.get("end"):
                    filter_info.append(f"日付範囲: {date_range['end']} 以前")
            
            if "tester" in filters:
                tester = filters["tester"]
                match_type = "完全一致" if tester.get("exact_match") else "部分一致"
                filter_info.append(f"担当者: {tester['value']} ({match_type})")
            
            if "result_type" in filters:
                filter_info.append(f"結果タイプ: {', '.join(filters['result_type'])}")
            
            if "environment" in filters:
                env = filters["environment"]
                match_type = "完全一致" if env.get("exact_match") else "部分一致"
                filter_info.append(f"環境: {env['value']} ({match_type})")
            
            if filter_info:
                metadata.append(["フィルタ条件", "; ".join(filter_info)])
        
        metadata.append(["設定ファイル", "config.json"])
        
        self._write_sheet_with_formatting(ws, metadata)
    
    def _write_sheet_with_formatting(self, ws, data):
        """シートにデータを書き込み、書式を適用"""
        # データ書き込み
        for row_idx, row in enumerate(data, 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 書式設定
        if data:
            # 游ゴシックフォント
            yugo_font = Font(name="游ゴシック", size=11)
            header_font = Font(name="游ゴシック", bold=True, size=11)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # ヘッダー行の書式設定
            for col_idx in range(1, len(data[0]) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # データ行の游ゴシックフォント適用
            for row_idx in range(2, len(data) + 1):
                for col_idx in range(1, len(data[0]) + 1):
                    ws.cell(row=row_idx, column=col_idx).font = yugo_font
            
            # 罫線設定
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for row_idx in range(1, len(data) + 1):
                for col_idx in range(1, len(data[0]) + 1):
                    ws.cell(row=row_idx, column=col_idx).border = thin_border
            
            # 列幅の自動調整
            # テスト結果列（Pass, Fixed, Fail, Blocked, Suspend, N/A, Total）はデフォルト幅
            result_headers = {"Pass", "Fixed", "Fail", "Blocked", "Suspend", "N/A", "Total"}
            for col_idx in range(1, len(data[0]) + 1):
                header = str(data[0][col_idx-1]) if data[0][col_idx-1] is not None else ""
                if header in result_headers:
                    continue  # デフォルト幅
                max_length = 0
                for row_idx in range(1, len(data) + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 4
    
    def generate_output_filename(self, base_filename, filters=None):
        """フィルタ条件を含むファイル名を生成"""
        if not filters:
            return base_filename
        
        filename_parts = []
        
        # 基本ファイル名
        name, ext = os.path.splitext(base_filename)
        filename_parts.append(name)
        
        # 日付範囲
        if "date_range" in filters:
            date_range = filters["date_range"]
            if date_range.get("start") and date_range.get("end"):
                filename_parts.append(f"{date_range['start']}_to_{date_range['end']}")
            elif date_range.get("start"):
                filename_parts.append(f"{date_range['start']}_onwards")
            elif date_range.get("end"):
                filename_parts.append(f"up_to_{date_range['end']}")
        
        # 担当者
        if "tester" in filters:
            filename_parts.append(filters["tester"]["value"])
        
        # 結果タイプ
        if "result_type" in filters:
            filename_parts.append("_".join(filters["result_type"]))
        
        # 環境
        if "environment" in filters:
            filename_parts.append(filters["environment"]["value"])
        
        return "_".join(filename_parts) + ext 