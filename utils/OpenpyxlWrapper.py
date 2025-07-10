from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime

def load(file_path:str, auto_create:bool=False):
    # ブックを開く
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        # ファイルが存在しない場合、新規作成
        if auto_create:
            wb = Workbook()
        else:
            raise FileNotFoundError(f"Error: {file_path} が見つかりません。")
    except PermissionError:
        raise PermissionError(f"Error: '{file_path}' は他のプログラムによって開かれています。")
    return wb

def create_sheet(workbook, sheet_name:str, overwrite:bool=False):
    # 既存のデータシートがあれば削除
    if overwrite and sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    # 新しいデータシートを作成
    return workbook.create_sheet(title=sheet_name)


def get_sheet_by_name(workbook, sheet_name:str):
    return workbook[sheet_name]


def get_sheetnames_by_keyword(workbook, keyword:str):
    return [sheet for sheet in workbook.sheetnames if keyword in sheet]

def get_sheetnames_by_keywords(workbook, keywords: list, ignores: list = None) -> list:
    """キーワードに基づいてワークブックからシート名をフィルタリングする

    Args:
        workbook: 対象のワークブック
        keywords: 検索キーワードのリスト
        ignores: 除外キーワードのリスト（デフォルト: None）

    Returns:
        list: フィルタリングされたシート名のリスト
    """
    # キーワードが指定されていない場合は全シート名を返す
    if len(keywords) == 0:
        return workbook.sheetnames

    # 除外キーワードが未指定の場合は空リストを使用
    if ignores is None:
        ignores = []

    # シート名のフィルタリング
    filtered_sheets = [
        sheet_name 
        for sheet_name in workbook.sheetnames 
        if _should_include_sheet(sheet_name, keywords, ignores)
    ]

    return filtered_sheets

def _should_include_sheet(sheet_name: str, keywords: list, ignores: list) -> bool:
    """シートを含めるべきかどうかを判定する

    Args:
        sheet_name: 判定対象のシート名
        keywords: 検索キーワードのリスト
        ignores: 除外キーワードのリスト

    Returns:
        bool: シートを含める場合はTrue、除外する場合はFalse
    """
    # いずれかのキーワードが含まれており、
    # かつ除外キーワードが含まれていないシートを選択
    has_keyword = any(keyword in sheet_name for keyword in keywords)
    has_ignore = any(ignore in sheet_name for ignore in ignores)
    
    return has_keyword and not has_ignore

def find_row(sheet, search_col:str, search_str:str):
    try:
        # 列名
        col_num = column_index_from_string(search_col) - 1

        # 指定列をループして値を確認
        for row in sheet.iter_rows(min_col=1, max_col=1):
            cell = row[col_num]  # 指定列のセル
            if cell.value == search_str:  # 値が search_str のセル
                return cell.row
        return None
    except Exception as e:
        print(f"Error: {e}")


def get_row_values(sheet, row_num:int):
    return [cell.value for cell in sheet[row_num]]


def get_column_values(sheet, col_nums: list, header_row: int = 1, ignore_header=False):
    if ignore_header:
        header_row += 1
    return [[sheet.cell(row=i, column=col_num).value for col_num in col_nums] for i in range(header_row, sheet.max_row + 1)]


def get_columns_data(sheet, col_nums: list, header_row: int = 1, ignore_header=False):
    if ignore_header:
        header_row += 1
    data = [[cell.value.strftime('%Y-%m-%d') if isinstance(cell.value, datetime) else cell.value 
             for col_num in col_nums 
             for cell in [sheet.cell(row=i, column=col_num)]] 
            for i in range(header_row, sheet.max_row + 1)]
    return data

def get_cell_value(sheet, col:int, row:int, replace_newline=False):
    value = sheet.cell(row=row, column=col).value
    if replace_newline:
        if value and isinstance(value, str):  # 値が文字列の場合のみ変換
            return value.replace("\n", "_")
    else:
        return value
